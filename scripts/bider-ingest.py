#!/usr/bin/env python3
"""Разбор Excel-выгрузок из Bider в CSV, которые читает агент.

Колонки Bider заранее не известны и меняются от вкладки к вкладке, поэтому
скрипт не угадывает схему: он находит строку заголовков, показывает, что нашёл,
и складывает данные как есть. Приведение к своим названиям — задача агента,
который читает CSV вместе с profile.md кабинета.

    python3 scripts/bider-ingest.py --input export.xlsx --account own --inspect
    python3 scripts/bider-ingest.py --input export.xlsx --account own

Зависимость: openpyxl (pip install openpyxl).
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("нужен openpyxl: pip install openpyxl")

# Сколько верхних строк просматриваем в поисках заголовков. Выгрузки обычно
# начинаются с названия отчёта и периода, поэтому не первая строка.
HEADER_SCAN_ROWS = 15


def slugify(name: str) -> str:
    """Имя листа → безопасное имя файла. Кириллица сохраняется."""
    s = re.sub(r"[^\w\-]+", "-", (name or "").strip(), flags=re.UNICODE)
    return s.strip("-").lower() or "sheet"


def find_header_row(rows: list[list], scan: int = HEADER_SCAN_ROWS) -> int:
    """Индекс строки заголовков: больше всего непустых ячеек, и под ней есть данные.

    Возвращает 0, если ничего похожего не нашлось — тогда лист пишется как есть.
    """
    best_idx, best_filled = 0, 0
    for i, row in enumerate(rows[:scan]):
        filled = sum(1 for c in row if c is not None and str(c).strip())
        has_data_below = any(
            any(c is not None and str(c).strip() for c in r) for r in rows[i + 1 : i + 4]
        )
        if filled > best_filled and has_data_below:
            best_idx, best_filled = i, filled
    return best_idx


def read_sheet(ws) -> list[list]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def trim_trailing_empty(rows: list[list]) -> list[list]:
    while rows and not any(c is not None and str(c).strip() for c in rows[-1]):
        rows.pop()
    return rows


def cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()[:10] if isinstance(value, dt.date) else value.isoformat()
    return str(value)



# --- Профиль «По дням» -------------------------------------------------------
# Заголовки Bider содержат эмодзи со скин-тонами и вариационными селекторами,
# поэтому лист опознаётся по устойчивым словам, а колонки маппятся по позиции.
DAILY_MARKERS = ("Переходы", "Ср. цена", "Сум. заказы")

DAILY_COLUMNS = [
    "date", "visits", "cart_cr_pct", "carts", "order_cr_pct", "orders",
    "buyout_pct", "bought", "in_transit", "cancelled", "avg_price",
    "orders_sum", "drr", "ad_impressions", "ad_ctr_pct", "ad_clicks",
    "ad_cpc", "ad_cpm", "ad_spend", "ad_cart_cr_pct", "ad_carts",
    "ad_carts_related", "ad_orders", "ad_orders_related",
]

# Пробелы бывают неразрывные и тонкие; «~?~» Bider ставит там, где значение
# ещё не определено (заказ не доехал) — это не ноль, это отсутствие данных.
_SPACES = "\u00a0\u2009\u202f\u2007 "
UNKNOWN_MARKERS = {"~?~", "?", "-", "—", ""}


def is_daily_sheet(headers: list[str]) -> bool:
    joined = " ".join(headers)
    return headers and headers[0].strip().lower().startswith("день") \
        and all(m in joined for m in DAILY_MARKERS)


def parse_number(raw: str):
    """'4 900 ₽' → 4900.0, '  6  %' → 6.0, '~?~' → None."""
    if raw is None:
        return None
    t = str(raw).strip()
    if t in UNKNOWN_MARKERS:
        return None
    for ch in _SPACES:
        t = t.replace(ch, "")
    t = t.replace("₽", "").replace("%", "").replace(",", ".")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def parse_date(raw: str) -> str:
    t = str(raw).strip()[:10]
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y"):
        try:
            return dt.datetime.strptime(t, fmt).date().isoformat()
        except ValueError:
            continue
    return t


def normalize_daily(headers: list[str], data: list[list]) -> tuple[list[str], list[list]]:
    """Приводит лист «По дням» к каноническим колонкам и числам.

    Колонок в выгрузке может быть больше или меньше ожидаемого — берём столько,
    сколько есть, лишние складываем в extra_N, чтобы ничего не потерять.
    """
    width = max((len(r) for r in data), default=len(headers))
    names = list(DAILY_COLUMNS[:width])
    names += [f"extra_{i}" for i in range(len(names) + 1, width + 1)]

    out = []
    for row in data:
        cells = list(row) + [None] * (width - len(row))
        rec = {"date": parse_date(cell_to_text(cells[0]))}
        for idx in range(1, width):
            rec[names[idx]] = parse_number(cell_to_text(cells[idx]))
        # Сквозная конверсия и выручка на переход — то, чего в выгрузке нет,
        # а именно по ним видно, что происходит с карточкой.
        visits = rec.get("visits") or 0
        rec["visit_to_order_pct"] = round(rec["orders"] / visits * 100, 2) \
            if visits and rec.get("orders") is not None else None
        rec["revenue_per_visit"] = round(rec["orders_sum"] / visits, 2) \
            if visits and rec.get("orders_sum") is not None else None
        out.append(rec)

    cols = names + ["visit_to_order_pct", "revenue_per_visit"]
    rows = [[rec.get(c) for c in cols] for rec in out]
    return cols, rows


def main() -> int:
    ap = argparse.ArgumentParser(description="Разбор выгрузок Bider в CSV")
    ap.add_argument("--input", required=True, type=Path, help="файл .xlsx из Bider")
    ap.add_argument("--account", default="own", help="slug кабинета (каталог в data/)")
    ap.add_argument("--lab-root", type=Path, default=Path.home() / ".claude-lab")
    ap.add_argument("--agent", default="nomado-market")
    ap.add_argument("--sheet", action="append", help="только эти листы (можно несколько)")
    ap.add_argument("--header-row", type=int, help="номер строки заголовков (с 1), если автопоиск ошибся")
    ap.add_argument("--inspect", action="store_true", help="показать структуру и выйти")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not args.input.exists():
        sys.exit(f"файл не найден: {args.input}")

    wb = load_workbook(args.input, read_only=True, data_only=True)
    out_dir = args.lab_root / args.agent / ".claude" / "data" / args.account
    stamp = dt.date.today().isoformat()

    print(f"Bider ingest · {args.input.name}")
    print(f"  кабинет:  {args.account}")
    if not args.inspect:
        print(f"  назначение: {out_dir}")
    print()

    written = []
    for ws in wb.worksheets:
        if args.sheet and ws.title not in args.sheet:
            continue

        rows = trim_trailing_empty(read_sheet(ws))
        if not rows:
            print(f"  — {ws.title}: пусто, пропуск")
            continue

        hdr = args.header_row - 1 if args.header_row else find_header_row(rows)
        headers = [cell_to_text(c) for c in rows[hdr]]
        data = rows[hdr + 1 :]
        # Строки-разделители и итоговые пустышки в выгрузках — обычное дело.
        data = [r for r in data if any(c is not None and str(c).strip() for c in r)]

        print(f"  — {ws.title}: заголовки в строке {hdr + 1}, данных {len(data)}")
        if args.inspect:
            shown = [h for h in headers if h][:12]
            print(f"      колонки: {', '.join(shown)}"
                  + (f" … ещё {len([h for h in headers if h]) - len(shown)}"
                     if len([h for h in headers if h]) > len(shown) else ""))
            continue

        dst = out_dir / f"{slugify(ws.title)}-{stamp}.csv"
        if args.dry_run:
            print(f"      [dry-run] → {dst.name}")
            written.append((ws.title, dst, len(data)))
            continue

        dst.parent.mkdir(parents=True, exist_ok=True)
        with dst.open("w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(headers)
            for r in data:
                w.writerow([cell_to_text(c) for c in r])
        print(f"      → {dst.name}")
        written.append((ws.title, dst, len(data)))

        # Сырой CSV остаётся как есть; нормализованный — то, что читает агент.
        if is_daily_sheet(headers):
            cols, rows = normalize_daily(headers, data)
            norm = dst.with_name(dst.stem + "-normalized.csv")
            with norm.open("w", newline="", encoding="utf-8") as fh:
                w = csv.writer(fh)
                w.writerow(cols)
                w.writerows(rows)
            print(f"      → {norm.name} (профиль «По дням»)")
            written.append((f"{ws.title} (нормализованный)", norm, len(rows)))

    if args.inspect:
        print("\nЕсли строка заголовков определена неверно — передайте --header-row N.")
        return 0

    if not written:
        print("\nНичего не записано.")
        return 1

    if not args.dry_run:
        index = out_dir / "README.md"
        with index.open("w", encoding="utf-8") as fh:
            fh.write(f"# Выгрузки Bider · кабинет `{args.account}`\n\n")
            fh.write(f"Последний импорт: {stamp}, источник `{args.input.name}`.\n\n")
            fh.write("| Лист | Файл | Строк |\n|---|---|---|\n")
            for title, dst, n in written:
                fh.write(f"| {title} | `{dst.name}` | {n} |\n")
            fh.write("\nСтарые выгрузки не удаляются: по ним видна динамика.\n")
        print(f"\nИндекс: {index}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
